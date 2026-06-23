import re, io
from datetime import datetime, date, time
import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.models.import_log import ImportLog

MONTH_ID = {
    "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
    "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
    "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12,
}
VALID_CREWS = {"SSE", "HAULER", "TYRE", "WHEEL"}
VALID_CODES = {"SCM", "USM", "TSM", "TUM", "ICM"}


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("", "####", "None", "nan") or s.startswith("=") else s


def clean_num(val):
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def parse_date(raw):
    """Parse '01 JUNI  2026' from row 4 col C (IMMOCS format)"""
    if not raw:
        return None
    m = re.search(r"(\d{1,2})\s+([A-Z]+)\s+(\d{4})", str(raw).upper())
    if not m:
        return None
    mon = MONTH_ID.get(m.group(2))
    if not mon:
        return None
    return date(int(m.group(3)), mon, int(m.group(1)))


def parse_fui_date(raw):
    """Parse per-row date from FUI Dashboard format: dd/mm/yyyy, d/m/yyyy, dd/mm/yy"""
    if raw is None:
        return None
    # Excel serial number
    try:
        n = float(str(raw).strip())
        if n > 40000:
            from datetime import timedelta
            excel_epoch = date(1899, 12, 30)
            return excel_epoch + timedelta(days=n)
    except (ValueError, TypeError):
        pass

    s = str(raw).strip()
    # dd/mm/yyyy or d/m/yyyy
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # dd/mm/yy
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2})$", s)
    if m:
        y = int(m.group(3))
        return date(2000 + y, int(m.group(2)), int(m.group(1)))
    # yyyy-mm-dd
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def parse_start_breakdown(val):
    """Parse '28/03/2026 (08:00)' or '17/08/24 (06:00)' -> datetime"""
    if val is None:
        return None
    s = str(val).strip()
    if s.startswith("=") or s in ("", "####", "None"):
        return None
    # dd/mm/yyyy (HH:MM)
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})\s*\((\d{2}):(\d{2})\)", s)
    if m:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)),
                        int(m.group(4)), int(m.group(5)))
    # dd/mm/yy (HH:MM)
    m = re.match(r"(\d{2})/(\d{2})/(\d{2})\s*\((\d{2}):(\d{2})\)", s)
    if m:
        y = int(m.group(3))
        y += 1900 if y > 50 else 2000
        return datetime(y, int(m.group(2)), int(m.group(1)),
                        int(m.group(4)), int(m.group(5)))
    return None


def parse_time_val(val):
    """Parse '06:00:00', '6:00:00 AM', datetime, time -> time or None"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, time):
        return val
    s = str(val).strip()
    if s.startswith("=") or s in ("", "####", "None", "#"):
        return None
    # "3:29:00 PM" format
    m = re.match(r"(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM)", s, re.I)
    if m:
        h, mi, se = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if m.group(4).upper() == "PM" and h != 12:
            h += 12
        if m.group(4).upper() == "AM" and h == 12:
            h = 0
        return time(h, mi, se)
    m = re.match(r"(\d{1,2}):(\d{2}):(\d{2})", s)
    if m:
        return time(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def detect_format(ws) -> str:
    """Detect sheet format: 'fui' if row 1 col A has 'DATE', else 'immocs'"""
    try:
        first_rows = list(ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True))
        if first_rows:
            v = first_rows[0][0]
            if v and str(v).strip().upper() == "DATE":
                return "fui"
    except Exception:
        pass
    return "immocs"


def parse_sheet(ws, source_file: str) -> list[dict]:
    """Parse one sheet (IMMOCS/DBR format) -> list of row dicts"""
    rd = parse_date(ws.cell(4, 3).value)
    if not rd:
        return []

    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        raw_cn = row[1]  # col B (index 1)
        if raw_cn is None or str(raw_cn).strip() in ("", "###") or str(raw_cn).startswith("="):
            continue
        cn = str(raw_cn).strip()

        crew = clean(row[2])
        if crew and crew not in VALID_CREWS:
            crew = None

        code = clean(row[4])
        if code:
            code = code.upper()
            if code not in VALID_CODES:
                code = None

        # row_no from col A
        raw_rn = row[0]
        if isinstance(raw_rn, (int, float)):
            row_no = int(raw_rn)
        else:
            row_no = r - 8

        start_bd = parse_start_breakdown(row[7])
        start_bd_hour = parse_time_val(row[8])
        rfu_bour = parse_time_val(row[9])

        # MO number (col L — index 11)
        mo_raw = clean(row[11])
        if mo_raw and mo_raw.replace(".", "").replace("-", "").isdigit():
            mo_raw = str(int(float(mo_raw)))

        rows.append({
            "report_date_raw": clean(ws.cell(4, 3).value),
            "report_date": rd,
            "breakdown_start_date": start_bd.date() if start_bd else None,
            "row_no": row_no,
            "cn": cn,
            "crew_type": crew,
            "section": clean(row[2]),
            "trouble_description": clean(row[3]),
            "breakdown_code": code,
            "hm_start": clean(row[5]),
            "hour_meter": clean_num(row[5]),
            "location": clean(row[6]),
            "start_breakdown": clean(row[7]),
            "start_bd_hour": start_bd_hour,
            "rfu_bour": rfu_bour,
            "start_time": None,
            "finish_time": None,
            "total": None,
            "total_hours": None,
            "hours_breakdown": None,
            "wo_number": clean(row[10]),
            "notification_number": clean(row[12]),
            "mo_number": mo_raw,
            "action_taken": clean(row[13]),
            "mechanic": clean(row[14]),
            "gl": clean(row[15]),
            "source_file": source_file,
        })
    return rows


def parse_fui_sheet(ws, source_file: str) -> list[dict]:
    """
    Parse FUI Dashboard format:
    Row 1: Header (DATE | C/N | SECTION | Trouble Description | Code | HM Start | LOC | Start Break Down | Start Time | Finish Time | Total | WO | NOTIFICATION | ACTION | MECHANIC | GL)
    Row 2+: Data, date per-row in col A
    """
    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=2, max_col=16, values_only=True), start=2):
        raw_date = row[0]
        rd = parse_fui_date(raw_date)
        if not rd:
            continue

        cn = row[1]
        if cn is None or str(cn).strip() == "":
            continue
        cn = str(cn).strip()

        crew = clean(row[2])
        if crew and crew not in VALID_CREWS:
            crew = None

        code = clean(row[4])
        if code:
            code = code.upper()
            if code not in VALID_CODES:
                code = None

        start_bd = parse_start_breakdown(row[7])

        # WO number (col L — index 11) — also could serve as MO if numeric
        raw_wo = clean(row[11])
        mo_raw = None
        if raw_wo and raw_wo.replace(".", "").replace("-", "").isdigit():
            mo_raw = str(int(float(raw_wo)))

        rows.append({
            "report_date_raw": clean(row[0]),
            "report_date": rd,
            "breakdown_start_date": start_bd.date() if start_bd else None,
            "row_no": r - 1,
            "cn": cn,
            "crew_type": crew,
            "section": clean(row[2]),
            "trouble_description": clean(row[3]),
            "breakdown_code": code,
            "hm_start": clean(row[5]),
            "hour_meter": clean_num(row[5]),
            "location": clean(row[6]),
            "start_breakdown": clean(row[7]),
            "start_bd_hour": None,
            "rfu_bour": None,
            "start_time": clean(row[8]),
            "finish_time": clean(row[9]),
            "total": clean(row[10]),
            "total_hours": None,
            "hours_breakdown": None,
            "wo_number": raw_wo,
            "notification_number": clean(row[12]),
            "mo_number": mo_raw,
            "action_taken": clean(row[13]),
            "mechanic": clean(row[14]),
            "gl": clean(row[15]),
            "source_file": source_file,
        })
    return rows


def import_excel(
    file_path: str, file_hash: str, filename: str, db: Session
):
    # --- Check duplicate file ---
    existing = (
        db.query(ImportLog)
        .filter(ImportLog.file_hash == file_hash)
        .first()
    )
    if existing:
        return {
            "status": "duplicate_file",
            "filename": filename,
            "total_rows": existing.total_rows,
        }

    # --- Parse Excel sheet by sheet ---
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        fmt = detect_format(ws)
        if fmt == "immocs":
            all_rows.extend(parse_sheet(ws, filename))
        else:
            all_rows.extend(parse_fui_sheet(ws, filename))
    wb.close()

    if not all_rows:
        return {
            "status": "empty_file",
            "filename": filename,
            "total_rows": 0,
        }

    # --- UPSERT into database ---
    insert_sql = text("""
    INSERT INTO dbr.breakdown_history
    (
        report_date_raw, report_date, breakdown_start_date,
        row_no, cn, crew_type, section,
        trouble_description, breakdown_code,
        hm_start, hour_meter, location,
        start_breakdown, start_bd_hour, rfu_bour,
        start_time, finish_time, total,
        total_hours, hours_breakdown,
        wo_number, notification_number, mo_number,
        action_taken, mechanic, gl,
        source_file
    )
    VALUES
    (
        :report_date_raw, :report_date, :breakdown_start_date,
        :row_no, :cn, :crew_type, :section,
        :trouble_description, :breakdown_code,
        :hm_start, :hour_meter, :location,
        :start_breakdown, :start_bd_hour, :rfu_bour,
        :start_time, :finish_time, :total,
        :total_hours, :hours_breakdown,
        :wo_number, :notification_number, :mo_number,
        :action_taken, :mechanic, :gl,
        :source_file
    )
    ON CONFLICT (report_date, cn, trouble_description, row_no)
    DO UPDATE SET
        action_taken = EXCLUDED.action_taken,
        mechanic = EXCLUDED.mechanic,
        gl = EXCLUDED.gl,
        rfu_bour = EXCLUDED.rfu_bour,
        source_file = EXCLUDED.source_file
    """)

    inserted = 0
    updated = 0

    try:
        for r in all_rows:
            db.execute(insert_sql, r)
            if db.bind.dialect.name == "postgresql":
                # approximate: not exact per row, but sync
                pass
            inserted += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise

    # --- Save import log ---
    import_log = ImportLog(
        filename=filename,
        file_hash=file_hash,
        total_rows=len(all_rows),
        inserted_rows=len(all_rows),
        skipped_rows=0,
    )
    db.add(import_log)
    db.commit()

    return {
        "status": "success",
        "filename": filename,
        "total_rows": len(all_rows),
        "inserted": len(all_rows),
        "skipped": 0,
    }
